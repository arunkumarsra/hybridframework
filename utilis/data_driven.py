import openpyxl


class DataDriven:
    path = r'C:\Users\arun\PycharmProjects\hybrid\testdata\Datadriven.xlsx'

    @staticmethod
    def get_rows():
        book = openpyxl.load_workbook(DataDriven.path)
        sheet = book.active
        rows = sheet.max_row
        return rows

    @staticmethod
    def get_columns():
        book = openpyxl.load_workbook(DataDriven.path)
        sheet = book.active
        columns = sheet.max_column
        return columns

    @staticmethod
    def get_value(rows, columns):
        book = openpyxl.load_workbook(DataDriven.path)
        sheet = book.active
        return sheet.cell(row=rows, column=columns).value


for x in range(2,DataDriven.get_rows()+1):
    for y in range(1,DataDriven.get_columns()+1):
        print(DataDriven.get_value(x, y))
        break
    break

